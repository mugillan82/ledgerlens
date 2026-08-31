import os
import io
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

def format_currency_str(val):
    """Formats numeric values to Indian Rupee format string (e.g. ₹1,23,456.78)."""
    try:
        val = float(val)
        prefix = "-" if val < 0 else ""
        s = f"{abs(val):.2f}"
        parts = s.split('.')
        integer = parts[0]
        fraction = parts[1]
        
        if len(integer) <= 3:
            res = integer
        else:
            last_three = integer[-3:]
            remaining = integer[:-3]
            out = []
            while len(remaining) > 0:
                out.insert(0, remaining[-2:])
                remaining = remaining[:-2]
            res = ",".join(out) + "," + last_three
        return f"{prefix}INR {res}.{fraction}"
    except (ValueError, TypeError):
        return str(val)

def generate_excel_report(summary_data, df_matches, df_exceptions, df_errors=None, output_path=None):
    """
    Creates a professionally formatted multi-sheet Excel reconciliation audit workbook.
    Sheets:
      1. Summary (KPIs, Match Rates, Business Impact, Accuracy)
      2. All Matches (bank_txn_id, gateway_order_id, match_type, confidence, fee, explanation)
      3. Exceptions (full exception list with reasons and amounts)
      4. Model Errors (FP/FN with failure explanations, if available)
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # ==================== SHEET 1: SUMMARY ====================
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["LedgerLens Reconciliation Audit Summary"])
    ws_sum.cell(row=1, column=1).font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    ws_sum.append([f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_sum.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws_sum.append([])
    
    # Section 1: Overview & Match Breakdown
    ws_sum.append(["Reconciliation Match Breakdown", ""])
    ws_sum.cell(row=4, column=1).fill = section_fill
    ws_sum.cell(row=4, column=1).font = section_font
    ws_sum.cell(row=4, column=2).fill = section_fill
    
    total_bank = summary_data.get("total_bank", 0)
    total_gateway = summary_data.get("total_gateway", 0)
    exact_count = summary_data.get("exact_count", 0)
    ml_count = summary_data.get("ml_count", 0)
    bank_exc_count = summary_data.get("bank_exc_count", 0)
    gate_exc_count = summary_data.get("gate_exc_count", 0)
    matched_count = exact_count + ml_count
    
    breakdown_rows = [
        ("Total Bank Records Processed", total_bank),
        ("Total Gateway Settlement Records", total_gateway),
        ("Exact Matches (1:1 Date/Amount/Ref)", exact_count),
        ("ML Predictive Matches", ml_count),
        ("Total Matches Cleared", matched_count),
        ("Bank Exceptions (Unmatched)", bank_exc_count),
        ("Gateway Exceptions (Unmatched)", gate_exc_count),
        ("Bank Match Rate", f"{(matched_count/total_bank*100):.2f}%" if total_bank > 0 else "0.00%")
    ]
    for label, val in breakdown_rows:
        ws_sum.append([label, val])
        ws_sum.cell(row=ws_sum.max_row, column=1).font = regular_font
        ws_sum.cell(row=ws_sum.max_row, column=2).font = bold_font
        ws_sum.cell(row=ws_sum.max_row, column=1).border = thin_border
        ws_sum.cell(row=ws_sum.max_row, column=2).border = thin_border
        
    ws_sum.append([])
    
    # Section 2: Business Impact
    ws_sum.append(["Business Impact & Audit Savings", ""])
    curr_row = ws_sum.max_row
    ws_sum.cell(row=curr_row, column=1).fill = section_fill
    ws_sum.cell(row=curr_row, column=1).font = section_font
    ws_sum.cell(row=curr_row, column=2).fill = section_fill
    
    tot_val = summary_data.get("total_value_processed", 0.0)
    rec_val = summary_data.get("total_value_reconciled", 0.0)
    exc_val = summary_data.get("total_value_exceptions", 0.0)
    hrs_saved = summary_data.get("hours_saved", 0.0)
    
    impact_rows = [
        ("Total Transaction Volume (Bank)", tot_val, True),
        ("Total Auto-Reconciled Value", rec_val, True),
        ("Total Flagged Exception Value (Pending Review)", exc_val, True),
        ("Reconciliation Value Coverage", f"{(rec_val/tot_val*100):.2f}%" if tot_val > 0 else "0.00%", False),
        ("Estimated Audit Labor Saved", f"{hrs_saved:.2f} hours (at 3 min/record)", False)
    ]
    for label, val, is_curr in impact_rows:
        ws_sum.append([label, val])
        r = ws_sum.max_row
        ws_sum.cell(row=r, column=1).font = regular_font
        ws_sum.cell(row=r, column=2).font = bold_font
        ws_sum.cell(row=r, column=1).border = thin_border
        ws_sum.cell(row=r, column=2).border = thin_border
        if is_curr:
            ws_sum.cell(row=r, column=2).number_format = '"INR "#,##0.00'
            
    # Section 3: Ground Truth Accuracy (if present)
    if "accuracy" in summary_data:
        ws_sum.append([])
        ws_sum.append(["Ground Truth Machine Learning Validation", ""])
        curr_row = ws_sum.max_row
        ws_sum.cell(row=curr_row, column=1).fill = section_fill
        ws_sum.cell(row=curr_row, column=1).font = section_font
        ws_sum.cell(row=curr_row, column=2).fill = section_fill
        
        gt_rows = [
            ("Accuracy", f"{summary_data.get('accuracy', 0.0):.2f}%"),
            ("Precision", f"{summary_data.get('precision', 0.0):.2f}%"),
            ("Recall", f"{summary_data.get('recall', 0.0):.2f}%"),
            ("F1 Score", f"{summary_data.get('f1_score', 0.0):.4f}" if "f1_score" in summary_data else "N/A"),
            ("True Positives (Correct Matches)", summary_data.get("tp", 0)),
            ("True Negatives (Correct Exceptions)", summary_data.get("tn", 0)),
            ("False Positives (Mismatched Pairs)", summary_data.get("fp", 0)),
            ("False Negatives (Missed Matches)", summary_data.get("fn", 0)),
        ]
        for label, val in gt_rows:
            ws_sum.append([label, val])
            r = ws_sum.max_row
            ws_sum.cell(row=r, column=1).font = regular_font
            ws_sum.cell(row=r, column=2).font = bold_font
            ws_sum.cell(row=r, column=1).border = thin_border
            ws_sum.cell(row=r, column=2).border = thin_border

    # ==================== SHEET 2: ALL MATCHES ====================
    ws_matches = wb.create_sheet(title="All Matches")
    ws_matches.views.sheetView[0].showGridLines = True
    
    match_headers = ["Bank Txn ID", "Gateway Order ID", "Match Type", "Confidence (%)", "Gateway Fee", "AI Explanation"]
    ws_matches.append(match_headers)
    for col_num in range(1, len(match_headers) + 1):
        cell = ws_matches.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    if df_matches is not None and not df_matches.empty:
        for _, row in df_matches.iterrows():
            conf = row.get("confidence", 100.0)
            fee_disp = row.get("fee_display", "")
            if not fee_disp and "fee_pct" in row:
                fee_disp = f"{row.get('fee_pct', 0.0):.2f}%"
            ws_matches.append([
                str(row.get("bank_txn_id", "")),
                str(row.get("gateway_order_id", "")),
                str(row.get("match_type", "")),
                float(conf) if conf != "" else 100.0,
                str(fee_disp),
                str(row.get("explanation", ""))
            ])
            r = ws_matches.max_row
            for c in range(1, len(match_headers) + 1):
                ws_matches.cell(row=r, column=c).font = regular_font
                ws_matches.cell(row=r, column=c).border = thin_border
            ws_matches.cell(row=r, column=4).number_format = '0.00'

    # ==================== SHEET 3: EXCEPTIONS ====================
    ws_exc = wb.create_sheet(title="Exceptions")
    ws_exc.views.sheetView[0].showGridLines = True
    
    if df_exceptions is not None and not df_exceptions.empty:
        exc_headers = list(df_exceptions.columns)
    else:
        exc_headers = ["txn_id", "source", "amount", "date", "reason", "audit_recommendation"]
        
    ws_exc.append(exc_headers)
    for col_num in range(1, len(exc_headers) + 1):
        cell = ws_exc.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    if df_exceptions is not None and not df_exceptions.empty:
        for _, row in df_exceptions.iterrows():
            row_vals = []
            for col in exc_headers:
                v = row.get(col, "")
                row_vals.append(v)
            ws_exc.append(row_vals)
            r = ws_exc.max_row
            for c in range(1, len(exc_headers) + 1):
                cell = ws_exc.cell(row=r, column=c)
                cell.font = regular_font
                cell.border = thin_border
                col_name = str(exc_headers[c-1]).lower()
                if "amount" in col_name and isinstance(cell.value, (int, float)):
                    cell.number_format = '"INR "#,##0.00'

    # ==================== SHEET 4: MODEL ERRORS ====================
    if df_errors is not None and not df_errors.empty:
        ws_err = wb.create_sheet(title="Model Errors")
        ws_err.views.sheetView[0].showGridLines = True
        err_headers = ["Error Type", "Bank Txn ID", "Bank Date", "Bank Amount", "Predicted Match", "Correct GT Match", "Failure Explanation"]
        ws_err.append(err_headers)
        for col_num in range(1, len(err_headers) + 1):
            cell = ws_err.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for _, row in df_errors.iterrows():
            ws_err.append([
                str(row.get("error_type", "")),
                str(row.get("bank_txn_id", "")),
                str(row.get("bank_date", "")),
                float(row.get("bank_amount", 0.0)) if str(row.get("bank_amount", "")).replace(".","",1).isdigit() else str(row.get("bank_amount", "")),
                str(row.get("pipeline_gateway_id", "")),
                str(row.get("correct_gateway_id", "")),
                str(row.get("failure_explanation", ""))
            ])
            r = ws_err.max_row
            for c in range(1, len(err_headers) + 1):
                cell = ws_err.cell(row=r, column=c)
                cell.font = regular_font
                cell.border = thin_border
                if c == 4 and isinstance(cell.value, (int, float)):
                    cell.number_format = '"INR "#,##0.00'

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        line_len = max(len(l) for l in lines)
                        if line_len > max_len:
                            max_len = line_len
                except:
                    pass
            adjusted_width = min(max(max_len + 4, 12), 65)
            sheet.column_dimensions[col_letter].width = adjusted_width

    # Write output
    if output_path:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream.getvalue()

class NumberedCanvas(canvas.Canvas):
    """Adds professional footer with page numbers and timestamp."""
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#595959"))
        
        # Header (pages > 1)
        if self._pageNumber > 1:
            self.drawString(40, 760, "LedgerLens Reconciliation Audit Report")
            self.setStrokeColor(colors.HexColor("#CCCCCC"))
            self.setLineWidth(0.5)
            self.line(40, 752, 572, 752)
            
        # Footer (all pages)
        self.setStrokeColor(colors.HexColor("#CCCCCC"))
        self.setLineWidth(0.5)
        self.line(40, 45, 572, 45)
        
        footer_text = f"Confidential Financial Audit Document — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        self.drawString(40, 32, footer_text)
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(572, 32, page_str)
        self.restoreState()

def generate_pdf_report(summary_data, df_exceptions, output_path=None):
    """
    Creates a clean, professional-looking audit PDF report containing:
    1. Title & Header with Date and Metadata
    2. Summary of Reconciliation & Business Impact
    3. Formatted Table of All Exceptions with Audit Reasons
    """
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        leftMargin=40,
        rightMargin=40,
        topMargin=50,
        bottomMargin=60
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#595959'),
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=17,
        textColor=colors.HexColor('#1F4E79'),
        spaceBefore=12,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=13,
        textColor=colors.HexColor('#262626')
    )
    
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#262626')
    )
    
    cell_header = ParagraphStyle(
        'TableCellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white
    )
    
    story = []
    
    # 1. Title Page Header
    story.append(Paragraph("LedgerLens Reconciliation Audit Report", title_style))
    gen_time = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    story.append(Paragraph(f"<b>Reconciliation Engine:</b> AI Finance Controller | <b>Date Generated:</b> {gen_time}", subtitle_style))
    story.append(Spacer(1, 8))
    
    # 2. Executive Summary & Business Impact Section
    story.append(Paragraph("1. Executive Summary & Business Impact", h2_style))
    
    tot_val = summary_data.get("total_value_processed", 0.0)
    rec_val = summary_data.get("total_value_reconciled", 0.0)
    exc_val = summary_data.get("total_value_exceptions", 0.0)
    hrs_saved = summary_data.get("hours_saved", 0.0)
    
    total_bank = summary_data.get("total_bank", 0)
    exact_count = summary_data.get("exact_count", 0)
    ml_count = summary_data.get("ml_count", 0)
    matched_count = exact_count + ml_count
    bank_exc_count = summary_data.get("bank_exc_count", 0)
    
    rec_pct = (rec_val / tot_val * 100) if tot_val > 0 else 0.0
    match_pct = (matched_count / total_bank * 100) if total_bank > 0 else 0.0
    
    summary_text = (
        f"This audit report certifies that <b>{total_bank} bank transactions</b> were evaluated against gateway settlement ledgers. "
        f"A total of <b>{matched_count} matches</b> ({match_pct:.1f}%) were successfully cleared automatically, auto-reconciling "
        f"<b>{format_currency_str(rec_val)}</b> ({rec_pct:.2f}% of total volume). "
        f"Automated processing saved approximately <b>{hrs_saved:.2f} hours</b> of manual auditing labor. "
        f"A total of <b>{bank_exc_count} bank exceptions</b> amounting to <b>{format_currency_str(exc_val)}</b> were isolated for manual review."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 10))
    
    # Summary Metrics Table
    summary_table_data = [
        [Paragraph("<b>Metric</b>", cell_header), Paragraph("<b>Value</b>", cell_header), Paragraph("<b>Business Implication</b>", cell_header)],
        [Paragraph("Total Processed Volume", cell_style), Paragraph(format_currency_str(tot_val), cell_style), Paragraph("Gross bank transaction statement value", cell_style)],
        [Paragraph("Auto-Reconciled Value", cell_style), Paragraph(format_currency_str(rec_val), cell_style), Paragraph(f"{rec_pct:.1f}% cleared without human intervention", cell_style)],
        [Paragraph("Outstanding Exceptions", cell_style), Paragraph(format_currency_str(exc_val), cell_style), Paragraph(f"{bank_exc_count} items requiring physical auditor sign-off", cell_style)],
        [Paragraph("Total Cleared Matches", cell_style), Paragraph(f"{matched_count} ({exact_count} exact, {ml_count} ML)", cell_style), Paragraph(f"{match_pct:.1f}% automated match rate", cell_style)],
        [Paragraph("Audit Hours Saved", cell_style), Paragraph(f"{hrs_saved:.2f} hrs", cell_style), Paragraph("Calculated at 3 minutes per matched transaction", cell_style)],
    ]
    
    sum_t = Table(summary_table_data, colWidths=[150, 130, 252])
    sum_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F4F7')])
    ]))
    story.append(sum_t)
    story.append(Spacer(1, 15))
    
    # 3. Exceptions Table Section (Main Auditor Content)
    story.append(Paragraph("2. Detailed Exception Log (Auditor Action Required)", h2_style))
    story.append(Paragraph(
        "The following transactions failed automated matching rules and ML confidence thresholds. "
        "Each exception includes its assigned root-cause categorization and audit recommendation:",
        body_style
    ))
    story.append(Spacer(1, 8))
    
    if df_exceptions is not None and not df_exceptions.empty:
        # Build exception rows
        # Expected columns: txn_id, source, amount, date, reason, audit_recommendation
        exc_table_data = [
            [
                Paragraph("<b>ID / Ref</b>", cell_header),
                Paragraph("<b>Source</b>", cell_header),
                Paragraph("<b>Date</b>", cell_header),
                Paragraph("<b>Amount</b>", cell_header),
                Paragraph("<b>Reason & Recommendation</b>", cell_header)
            ]
        ]
        
        for _, row in df_exceptions.iterrows():
            t_id = str(row.get("txn_id", row.get("bank_txn_id", "")))
            src = str(row.get("source", "Bank"))
            dt = str(row.get("date", ""))
            amt = format_currency_str(row.get("amount", 0.0))
            reason = str(row.get("reason", "Unmatched transaction"))
            rec = str(row.get("audit_recommendation", ""))
            combined_reason = f"<b>Reason:</b> {reason}" + (f"<br/><b>Action:</b> {rec}" if rec else "")
            
            exc_table_data.append([
                Paragraph(t_id, cell_style),
                Paragraph(src, cell_style),
                Paragraph(dt, cell_style),
                Paragraph(amt, cell_style),
                Paragraph(combined_reason, cell_style)
            ])
            
        exc_t = Table(exc_table_data, colWidths=[85, 55, 65, 85, 242], repeatRows=1)
        exc_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
        ]))
        story.append(exc_t)
    else:
        story.append(Paragraph("<b>No exceptions found.</b> All transactional records were fully reconciled.", body_style))
        
    # Build Document
    doc.build(story, canvasmaker=NumberedCanvas)
    pdf_bytes = pdf_buffer.getvalue()
    pdf_buffer.close()
    
    if output_path:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
            
    return pdf_bytes
